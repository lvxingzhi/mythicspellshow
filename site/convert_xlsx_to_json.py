"""
将 xlsx/ 目录下的所有 Excel 文件转换为前端查询站使用的 data.json。

支持：
    - 自动扫描 xlsx/ 目录下所有 .xlsx 文件
    - 按文件名前缀分类赛季（如 Midnight_xxx.xlsx → Midnight）
    - 同一副本内按 (npcId, spellId) 去重

用法:
    python convert_xlsx_to_json.py                  # 默认读取 ../xlsx/，输出 data.json
    python convert_xlsx_to_json.py <xlsx_dir> [output.json]

依赖:
    pip install openpyxl
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


# xlsx 中的中文列名 → 前端字段名
COLUMN_MAP = {
    "副本名(中)": "dungeonZh",
    "副本名(英)": "dungeonEn",
    "怪物名(中)": "mobZh",
    "怪物名(英)": "mobEn",
    "NPC ID": "npcId",
    "生物类型": "creatureType",
    "控制技能": "ccTypes",
    "是否BOSS": "isBoss",
    "技能名": "spellName",
    "Spell ID": "spellId",
    "技能描述": "description",
    "可打断": "interruptible",
    "魔法效果": "isMagic",
    "激怒": "isEnrage",
    "流血": "isBleed",
    "中毒": "isPoison",
    "疾病": "isDisease",
    "诅咒": "isCurse",
    
}

# 文件名格式即为赛季名，如 12.0-S1.xlsx → 赛季 12.0-S1
SEASON_FROM_FILENAME = re.compile(r"^(.+)\.xlsx$")


def to_bool(value: object) -> bool:
    """把 '是 / 否 / True / 1' 之类的混合值规范化为 bool。"""
    if value is None:
        return False
    s = str(value).strip().lower()
    return s in {"是", "true", "1", "yes", "y"}


def extract_season(filename: str) -> str:
    """从文件名提取赛季名，如 12.0-S1.xlsx → 12.0-S1。"""
    m = SEASON_FROM_FILENAME.match(filename)
    return m.group(1) if m else "unknown"


def parse_xlsx(xlsx_path: Path, season: str) -> tuple[list[dict], list[dict]]:
    """解析单个 xlsx 文件，返回 (dungeons, rows)。

    Args:
        xlsx_path: Excel 文件路径
        season: 赛季名（如 12.0-S1）

    Returns:
        dungeons: 副本摘要列表
        rows: 所有数据行（每行附带 season 字段）
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    dungeons: list[dict] = []
    all_rows: list[dict] = []

    for sheet in wb.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        idx = {name: i for i, name in enumerate(header) if name in COLUMN_MAP}
        if not idx:
            continue

        sheet_rows: list[dict] = []
        for row in rows_iter:
            if not row or all(c is None or c == "" for c in row):
                continue
            item: dict = {}
            for cn, en in COLUMN_MAP.items():
                if cn not in idx:
                    continue
                val = row[idx[cn]]
                if en in ("interruptible", "isMagic", "isBoss", "isEnrage", "isBleed", "isPoison", "isDisease", "isCurse"):
                    item[en] = to_bool(val)
                else:
                    item[en] = "" if val is None else str(val).strip()
            sheet_rows.append(item)

        if sheet_rows:
            # 给每行附上赛季标识
            for r in sheet_rows:
                r["season"] = season
            dungeon_name = sheet_rows[0].get("dungeonZh") or sheet.title
            dungeons.append({"name": dungeon_name, "sheet": sheet.title, "count": len(sheet_rows)})
            all_rows.extend(sheet_rows)

    wb.close()
    return dungeons, all_rows


def deduplicate(rows: list[dict]) -> list[dict]:
    """按 (npcId, spellId) 去重，保留最后出现的记录。

    不同副本出现相同的 npcId+spellId 组合时，保留各自记录；
    同一副本内重复的 npcId+spellId 只保留一条。

    Args:
        rows: 原始数据行

    Returns:
        去重后的数据行
    """
    seen: dict[tuple[str, str, str, str], dict] = {}  # (season, dungeonZh, npcId, spellId) → row
    for row in rows:
        key = (row.get("season", ""), row.get("dungeonZh", ""), row.get("npcId", ""), row.get("spellId", ""))
        seen[key] = row  # 后出现的覆盖前面的
    return list(seen.values())


def convert(xlsx_dir: Path, json_path: Path) -> None:
    """扫描 xlsx 目录，解析所有 Excel 文件并合并去重，输出 data.json。

    Args:
        xlsx_dir: 存放 xlsx 文件的目录
        json_path: 输出 json 文件路径
    """
    xlsx_files = sorted(xlsx_dir.glob("*.xlsx"))
    if not xlsx_files:
        print(f"错误: {xlsx_dir} 下未找到任何 .xlsx 文件")
        sys.exit(1)

    all_dungeons: list[dict] = []
    all_rows: list[dict] = []
    seasons: dict[str, list[str]] = {}  # season → [filename, ...]

    for xlsx_path in xlsx_files:
        season = extract_season(xlsx_path.name)
        seasons.setdefault(season, []).append(xlsx_path.name)

        print(f"  解析: {xlsx_path.name} (赛季: {season})")
        dungeons, rows = parse_xlsx(xlsx_path, season)
        all_dungeons.extend(dungeons)
        all_rows.extend(rows)
        print(f"    {len(dungeons)} 个副本, {len(rows)} 行数据")

    # 去重
    before = len(all_rows)
    all_rows = deduplicate(all_rows)
    after = len(all_rows)
    if before != after:
        print(f"  去重: {before} → {after} 行（移除 {before - after} 条重复）")

    # 副本列表也去重（按 name 保留最后出现的）
    dungeon_map: dict[str, dict] = {}
    for d in all_dungeons:
        dungeon_map[d["name"]] = d
    all_dungeons = list(dungeon_map.values())
    # 重新统计去重后的行数
    for d in all_dungeons:
        d["count"] = sum(1 for r in all_rows if r.get("dungeonZh") == d["name"])

    payload = {
        "seasons": seasons,
        "dungeons": all_dungeons,
        "rows": all_rows,
        "total": len(all_rows),
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n已写出 {json_path} （共 {len(all_rows)} 行，{len(all_dungeons)} 个副本，{len(seasons)} 个赛季）")


def main() -> None:
    xlsx_dir = Path(sys.argv[1]).expanduser().resolve() if len(sys.argv) >= 2 else Path(__file__).resolve().parent.parent / "xlsx"
    json_path = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) >= 3 else Path(__file__).resolve().parent / "data.json"

    if not xlsx_dir.is_dir():
        print(f"错误: {xlsx_dir} 不是有效目录")
        sys.exit(1)

    print(f"扫描目录: {xlsx_dir}")
    convert(xlsx_dir, json_path)


if __name__ == "__main__":
    main()